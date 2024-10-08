#!/usr/bin/env python3
from collections import namedtuple
from json import load, loads
from base64 import b64decode
from sys import stdin
from argparse import ArgumentParser
from openpyxl import Workbook
from report_lib import *
from statistics import mean, median, pstdev

host_report = dict()

for k, v in load(stdin).items():
    if v["retcode"] == 0:
        host_report[k] = loads(b64decode(v["stdout"]).decode())

report = {
    "mem_all_hosts": {
        "mem_total": 0.0,
        "mem_available": 0.0,
        "ratio": 0.0
    },
    "top_n_hosts_least_mem_available": list(),
    "top_n_procs_biggest_rss": list(),
    "mem_all_vms": {
        "memory": 0.0,
        "rss": 0.0,
        "ratio": 0.0
    },
    "top_n_vms_by_memory": list(),
    "top_n_vms_by_rss": list(),
    "top_n_vms_by_ratio": list(),
    "proc_statistics": list(),
    "host_statistics": list()
}

hosts = list()
procs = list()
vms = list()
os_ram_allocation_ratio_global = 1.5
Host = namedtuple("Host", ["host", "mem_total", "mem_available", "ratio", "os_ram_allocation_ratio", "os_reserved_host_memory_mb"])
Proc = namedtuple("Proc", ["host", "name", "rss"])
Vm = namedtuple("Vm", ["host", "uuid", "memory", "rss", "ratio"])
ProcS = namedtuple("ProcS", ["name", "mean_rss", "median_rss", "max_rss", "pstdev_rss", "pstdev_rss_ratio"])    # Proc statistics
HostS = namedtuple("HostS", ["host", "mem_total", "mem_available", "mem_procs_rss", "mem_vms_allocated", "mem_vms_rss", "os_ram_allocation_ratio", "os_reserved_host_memory_mb"])   # Host statistics

for hk, hv in host_report.items():
    host = hk.split(".")[0]
    if host.lower().startswith("ctl0"):
        if hv["host"]["os_ram_allocation_ratio"] != 0.0:
            os_ram_allocation_ratio_global = hv["host"]["os_ram_allocation_ratio"]
    else:
        report["mem_all_hosts"]["mem_total"] += hv["host"]["mem_total"]
        report["mem_all_hosts"]["mem_available"] += hv["host"]["mem_available"]
        hosts.append(Host(host, hv["host"]["mem_total"], hv["host"]["mem_available"], hv["host"]["ratio"], hv["host"]["os_ram_allocation_ratio"], hv["host"]["os_reserved_host_memory_mb"]))
        host_statistics = {
            "mem_total": hosts[-1].mem_total,
            "mem_available": hosts[-1].mem_available,
            "mem_procs_rss": 0.0,
            "mem_vms_allocated": 0,
            "mem_vms_rss": 0.0,
            "os_ram_allocation_ratio": hosts[-1].os_ram_allocation_ratio,
            "os_reserved_host_memory_mb": hosts[-1].os_reserved_host_memory_mb
        }
        for pk, pv in hv["proc"].items():
            procs.append(Proc(host, pv["name"], pv["rss"]))
            host_statistics["mem_procs_rss"] += procs[-1].rss
        for vk, vv in hv["vm"].items():
            report["mem_all_vms"]["memory"] += vv["memory"]
            report["mem_all_vms"]["rss"] += vv["rss"]
            vms.append(Vm(host, vk, vv["memory"], vv["rss"], vv["ratio"]))
            host_statistics["mem_vms_allocated"] += vms[-1].memory
            host_statistics["mem_vms_rss"] += vms[-1].rss
        report["host_statistics"].append(HostS(host, **host_statistics))
for i in range(len(hosts)):
    if hosts[i].os_ram_allocation_ratio == 0.0:
        hosts[i] = hosts[i]._replace(os_ram_allocation_ratio=os_ram_allocation_ratio_global)
for i in range(len(report["host_statistics"])):
    if report["host_statistics"][i].os_ram_allocation_ratio == 0.0:
        report["host_statistics"][i] = report["host_statistics"][i]._replace(os_ram_allocation_ratio=os_ram_allocation_ratio_global)
report["host_statistics"] = sorted(report["host_statistics"], key=lambda a: a.mem_available)

report["mem_all_hosts"]["ratio"] = round(report["mem_all_hosts"]["mem_available"] / report["mem_all_hosts"]["mem_total"], 2) if report["mem_all_hosts"]["mem_total"] != 0 else 0
report["mem_all_vms"]["ratio"] = round(report["mem_all_vms"]["rss"] / report["mem_all_vms"]["memory"], 2) if report["mem_all_vms"]["memory"] != 0 else 0
report["top_n_hosts_least_mem_available"] = sorted(hosts, key=lambda a: a.mem_available)
report["top_n_procs_biggest_rss"] = sorted(procs, key=lambda a: a.rss, reverse=True)
report["top_n_vms_by_memory"] = sorted(vms, key=lambda a: a.memory, reverse=True)
report["top_n_vms_by_rss"] = sorted(vms, key=lambda a: a.rss, reverse=True)
report["top_n_vms_by_ratio"] = sorted(vms, key=lambda a: a.ratio, reverse=True)

proc_statistics = dict()
for p in procs:
    if p.name not in proc_statistics:
        proc_statistics[p.name] = {
            "rss": [p.rss]
        }
    else:
        proc_statistics[p.name]["rss"].append(p.rss)
for k, v in proc_statistics.items():
    mean_rss = mean(v["rss"])
    pstdev_rss = pstdev(v["rss"])
    pstdev_rss_ratio = pstdev_rss / mean_rss
    report["proc_statistics"].append(ProcS(k, round(mean_rss, 2), round(median(v["rss"]), 2), round(max(v["rss"]), 2), round(pstdev_rss, 2), round(pstdev_rss_ratio, 2)))
report["proc_statistics"] = sorted(report["proc_statistics"], key=lambda a: a.median_rss, reverse=True)

parser = ArgumentParser(description="Merge reports from hosts")
parser.add_argument("-n", metavar="N", type=int, required=False, default=10, help="Top N-items to report")
parser.add_argument("-x", default=False, required=False, action="store_true", help="Write Excel report")
args = parser.parse_args()

n = args.n

print("Total hosts/procs/vms")
print("hosts {}, procs {}, vms {}".format(len(hosts), len(procs), len(vms)))
print("")
print("Total/available memory across all hosts")
print("mem_total {mem_total:.2f}, mem_available {mem_available:.2f}, ratio {ratio:.2f}".format(**report["mem_all_hosts"]))
print("")
print("Total memory/rss of all VMs across all hosts")
print("memory {memory:.2f}, rss {rss:.2f}, ratio {ratio:.2f}".format(**report["mem_all_vms"]))
print("")
print("Top {} hosts with least memory available".format(n))
for i in report["top_n_hosts_least_mem_available"][:n]:
    print(i)
print("")
print("Top {} processes with biggest rss".format(n))
for i in report["top_n_procs_biggest_rss"][:n]:
    print(i)
print("")

if args.x:
    wb = Workbook()

    ws = wb.active
    ws.title = "Totals"
    ws.append(["Total hosts/procs/vms", "hosts", "procs", "vms"])
    ws.append([None, len(hosts), len(procs), len(vms)])
    ws.append(["Total/available memory across all hosts", "mem_total", "mem_available", "ratio"])
    ws.append([None] + [round(report["mem_all_hosts"][x], 2) for x in ["mem_total", "mem_available", "ratio"]])
    ws.append(["Total memory/rss of all VMs across all hosts", "memory", "rss", "ratio"])
    ws.append([None] + [round(report["mem_all_vms"][x], 2) for x in ["memory", "rss", "ratio"]])
    cell_props = CellProperties(ws[1][0])
    for row_i in range(1, 6, 2):
        for cell in ws[row_i]:
            cell_props.HeaderFont(cell)

    reports_meta = [
        {
            "name": "top_n_hosts_least_mem_available",
            "header": "Top {} hosts with least memory available".format(n),
            "top": True
        },
        {
            "name": "top_n_procs_biggest_rss",
            "header": "Top {} processes with biggest rss".format(n),
            "top": True
        },
        {
            "name": "top_n_vms_by_memory",
            "header": "Top {} VMs by memory".format(n),
            "top": True
        },
        {
            "name": "top_n_vms_by_rss",
            "header": "Top {} VMs by rss".format(n),
            "top": True
        },
        {
            "name": "top_n_vms_by_ratio",
            "header": "Top {} VMs by ratio".format(n),
            "top": True
        },
        {
            "name": "proc_statistics",
            "header": "Proc statistics",
            "top": False
        },
        {
            "name": "host_statistics",
            "header": "Host statistics",
            "top": False
        }
    ]
    for r in reports_meta:
        ws = wb.create_sheet(r["name"])
        ws.append((r["header"],) + report[r["name"]][0]._fields)
        boundary = len(report[r["name"]])
        if r["top"]:
            boundary = n
        for i in report[r["name"]][:boundary]:
            ws.append(("",) + tuple(i))
        for cell in ws[1]:
            cell_props.HeaderFont(cell)
        ws.freeze_panes = "B2"

    for ws in wb.worksheets:
        for row_i in range(1, ws.max_row + 1):
            adjust_col_width(ws, row_i)

    report_file = prepare_report_file()
    logmsg("Write report {}".format(report_file))
    wb.save(report_file)
