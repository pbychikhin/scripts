
from concurrent.futures import ThreadPoolExecutor
from fractions import Fraction

# __main__-related imports
# import json
from openpyxl import Workbook
from openpyxl.comments import Comment
from copy import copy
from os.path import exists
from os.path import join as path_join

# common lib
from report import *


def os_api_get_loadbalancers(auth_obj: OsAuth, url, page_sz=0, num_pages=0):
    return os_api_get_resources_gen(auth_obj, url, "lbaas/loadbalancers", "loadbalancers", "loadbalancers_links", page_sz, num_pages)


def os_api_get_members(auth_obj: OsAuth, url, pool_id, page_sz=0, num_pages=0):
    return os_api_get_resources_gen(auth_obj, url, "/".join(("lbaas/pools", pool_id, "members")), "members", "members_links", page_sz, num_pages)


def os_api_get_loadbalancer(auth_obj: OsAuth, url, lb_id):
    return os_api_get(auth_obj, url, "/".join(("lbaas/loadbalancers", lb_id))).json()["loadbalancer"]


def os_api_get_loadbalancer_status(auth_obj: OsAuth, url, lb_id):
    status = os_api_get(auth_obj, url, "/".join(("lbaas/loadbalancers", lb_id, "status"))).json()["statuses"]["loadbalancer"]
    for listener in status["listeners"]:
        for pool in listener["pools"]:
            pool["members"] = [x for x in os_api_get_members(auth_obj, url, pool["id"])]
    return status


def os_api_get_loadbalancers_with_status(auth_obj: OsAuth, url):
    """
    Combines loadbalancer {details} and {status} three into a single dict
    :param auth_obj: Openstack auth obj
    :param url: Explicit URL
    :return: Dict of loadbalancers
    """
    rv = dict()
    futures = list()
    with ThreadPoolExecutor(max_workers=10) as executor:
        for i in os_api_get_loadbalancers(auth_obj, url):
            rv[i["id"]] = dict()
            rv[i["id"]]["details"] = i
            futures.append({"id": i["id"], "future": executor.submit(os_api_get_loadbalancer_status, auth_obj, url, i["id"])})
    for i in futures:
        if i["future"].exception():
            print("LB id {}, exception getting status: {}".format(i["id"], i["future"].exception()), file=sys.stderr)
        else:
            rv[i["id"]]["status"] = i["future"].result()
    return rv


def os_api_get_amphorae(auth_obj: OsAuth, url, page_sz=0, num_pages=0):
    """
    Returns list of amphorae per LB
    :param auth_obj: Openstack auth obj
    :param url: Explicit URL
    :param page_sz: Result pagination limit (positive integer or 0 to disable limit)
    :param num_pages: Number of pages to request (positive integer or 0 to disable limit)
    :return: List of amphorae per LB id
    """
    rv = dict()
    for i in os_api_get_resources_gen(auth_obj, url, "octavia/amphorae", "amphorae", "amphorae_links", page_sz, num_pages):
        if i["loadbalancer_id"] not in rv:
            rv[i["loadbalancer_id"]] = list()
        rv[i["loadbalancer_id"]].append(i)
    return rv


def inspect_amphorae(amphorae, ports: OsPorts=None):
    """
    Inspects properties of amphorae of a balancer and returns insight as text notes
    :param amphorae: list loadbalancer's amphorae
    :param ports: OsPorts obj
    :return: inspections rv obj
    """
    inspections = OrderedDict((
        (
            # Amphora not found
            "NOT_FOUND",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            "vrrp_port_NOT_FOUND",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            "ha_port_NOT_FOUND",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # Amphora status is ERROR
            "ERROR",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # vrrp_port status is ERROR
            "vrrp_port_ERROR",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # ha_port status is ERROR
            "ha_port_ERROR",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # Amphora status is not ALLOCATED
            "not_ALLOCATED_or_ACTIVE",
            {
                "severity": InspectionSeverity.MEDIUM,
                "count": 0
            }
        ),
        (
            # vrrp_port status is not ACTIVE
            "vrrp_port_not_ACTIVE",
            {
                "severity": InspectionSeverity.MEDIUM,
                "count": 0
            }
        ),
        (
            # ha_port status is not ADMIN_DOWN (ADMIN_DOWN seems normal for ha_port)
            "ha_port_not_ADMIN_DOWN",
            {
                "severity": InspectionSeverity.MEDIUM,
                "count": 0
            }
        ),
        (
            # Amphora is Ok
            "Ok",
            {
                "severity": InspectionSeverity.LOW,
                "count": 1
            }
        )
    ))

    count = 0
    rv = list()
    if amphorae:
        if len(amphorae) > 1:
            rv_template = "A{count}: status: {status}"
        else:
            rv_template = "status: {status}"
        for amphora in amphorae:
            port_str = {"vrrp_port": "", "ha_port": ""}
            if ports:
                rv_template = ", ".join((rv_template, "vrrp_port: {vrrp_port}, ha_port: {ha_port}"))
                for k in port_str.keys():
                    try:
                        port = ports.get_by_id(amphora["{}_id".format(k)])
                        if port["status"] == "DOWN" and not port["admin_state_up"]:
                            port_str[k] = "ADMIN_DOWN"
                            if k == "vrrp_port":
                                inspections["vrrp_port_not_ACTIVE"]["count"] += 1
                        else:
                            port_str[k] = port["status"]
                            if port_str[k] == "ERROR":
                                inspections["{}_ERROR".format(k)]["count"] += 1
                            elif k == "ha_port":
                                inspections["{}_not_ADMIN_DOWN".format(k)]["count"] += 1
                    except KeyError:
                        port_str[k] = "NOT_FOUND"
                        inspections["{}_NOT_FOUND".format(k)]["count"] += 1
            rv.append(rv_template.format(count=count, status=amphora["status"], **port_str))
            count += 1
            if amphora["status"] == "ERROR":
                inspections["ERROR"]["count"] += 1
            elif amphora["status"] not in ("ALLOCATED", "ACTIVE"):
                inspections["not_ALLOCATED_or_ACTIVE"]["count"] += 1
    else:
        rv.append("NOT_FOUND")
        inspections["NOT_FOUND"]["count"] += 1
    return make_inspection_rv(inspections, rv)


def inspect_members(loadbalancer, ports: OsPorts=None):
    """
    Inspects properties of members of all pools of a loadbalancer and returns insight as text notes
    :param loadbalancer: loadbalancer obj
    :param ports: OsPorts obj
    :return: inspections rv obj
    """
    inspections = OrderedDict((
        (
            # No listeners found
            "NO_LISTENERS_FOUND",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # No members found
            "NOT_FOUND",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # Ports not found for threshold_ratio or more members
            "PORTS_NOT_FOUND",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # Provisioning status is ERROR for threshold_ratio or more members
            "provisioning_ERROR",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # Operating status is ERROR for threshold_ratio or more members
            "operating_ERROR",
            {
                "severity": InspectionSeverity.HIGH,
                "count": 0
            }
        ),
        (
            # Provisioning status is not ACTIVE for threshold_ratio or more members
            "provisioning_not_ACTIVE",
            {
                "severity": InspectionSeverity.MEDIUM,
                "count": 0
            }
        ),
        (
            # Operating status is not ONLINE for threshold_ratio or more members
            "operating_not_ONLINE",
            {
                "severity": InspectionSeverity.MEDIUM,
                "count": 0
            }
        ),
        (
            # Members are Ok
            "Ok",
            {
                "severity": InspectionSeverity.LOW,
                "count": 1
            }
        )
    ))

    threshold_ratio = Fraction(2, 3)
    total = 0
    statuses = {"provisioning": dict(), "operating": dict()}
    combined_statuses = dict()
    ports_not_found = 0
    rv = list()
    listeners = loadbalancer["status"]["listeners"]
    if listeners:
        for listener in listeners:
            for pool in listener["pools"]:
                for member in pool["members"]:
                    total += 1
                    for key in ("provisioning", "operating"):
                        member_key = "{}_status".format(key)
                        if member[member_key] not in statuses[key]:
                            statuses[key][member[member_key]] = 0
                        statuses[key][member[member_key]] += 1
                    combined_status = "{provisioning_status}/{operating_status}".format(**member)
                    if combined_status not in combined_statuses:
                        combined_statuses[combined_status] = 0
                    combined_statuses[combined_status] += 1
                    if ports:
                        if not ports.get_by_subnet_ip(str(member["subnet_id"] or loadbalancer["details"]["vip_subnet_id"]), str(member["address"])):
                            ports_not_found += 1
        rv.append("TOTAL: {}".format(total))
        rv.extend("{}: {}".format(x, combined_statuses[x]) for x in sorted(combined_statuses.keys()))
        if ports_not_found:
            rv.append("PORTS_NOT_FOUND: {}".format(ports_not_found))
            if total > 0 and Fraction(ports_not_found, total) >= threshold_ratio:
                inspections["PORTS_NOT_FOUND"]["count"] += 1
        for k, v in statuses.items():
            error_key, not_good_key = None, None
            if k == "provisioning":
                error_key = "provisioning_ERROR"
                not_good_key = "provisioning_not_ACTIVE"
            if k == "operating":
                error_key = "operating_ERROR"
                not_good_key = "operating_not_ONLINE"
            error_count = 0
            not_good_count = 0
            for kk, vv in v.items():
                if kk == "ERROR":
                    error_count += vv
                if k == "provisioning":
                    if kk != "ACTIVE":
                        not_good_count += vv
                if k == "operating":
                    if kk not in ("ONLINE", "NO_MONITOR"):
                        not_good_count += vv
            for _key, _count in zip((error_key, not_good_key), (error_count, not_good_count)):
                if total > 0 and Fraction(_count, total) >= threshold_ratio:
                    inspections[_key]["count"] += 1
    else:
        rv.append("NO_LISTENERS_FOUND")
        inspections["NO_LISTENERS_FOUND"]["count"] += 1
    if total == 0:
        inspections["NOT_FOUND"]["count"] += 1
    return make_inspection_rv(inspections, rv)


def inspect_loadbalancer(loadbalancer, ports:OsPorts):
    """
    Inspects the loadbalancer itself to fill up comments for id column
    :param loadbalancer: loadbalancer obj
    :param ports: ports obj
    :return: inspections rv obj
    """
    inspections = OrderedDict((
        (
            # Some warnings about LB
            "WARNING",
            {
                "severity": InspectionSeverity.MEDIUM,
                "count": 0
            }
        ),
        (
            # LB is Ok
            "Ok",
            {
                "severity": InspectionSeverity.LOW,
                "count": 1
            }
        )
    ))

    port_owners = dict()
    rv = list()
    for port in ports.get_by_network(loadbalancer["details"]["vip_network_id"]):
        device_owner = port["device_owner"]
        if device_owner.strip() == "":
            device_owner = "NO_OWNER"
        elif device_owner == "compute:nova" and port["name"].startswith("octavia-lb-"):
            device_owner = "Octavia"
        elif device_owner.startswith("network:"):
            device_owner = "network"
        elif device_owner.startswith("compute:"):
            device_owner = "compute"
        if device_owner not in port_owners:
            port_owners[device_owner] = 0
        port_owners[device_owner] += 1
    for k in sorted(port_owners.keys()):
        rv.append("{} devs: {}".format(k, port_owners[k]))
    if rv:
        rv = [", ".join(rv)]
    if not port_owners or set(port_owners.keys()) <= {"Octavia", "network", "NO_OWNER"}:
        inspections["WARNING"]["count"] += 1
        rv.insert(0, "EMPTY_VIP_NETWORK")
    return make_inspection_rv(inspections, rv)


if __name__ == "__main__":
    logmsg("Auth")
    auth = OsAuth()

    logmsg("Get LB URL")
    loadbalancer_url = os_api_get_service_url(auth, "load-balancer")

    logmsg("Get Network URL")
    network_url = os_api_get_service_url(auth, "network")

    logmsg("Get projects")
    projects = os_api_get_projects(auth, auth.identity_url)

    logmsg("Get amphorae")
    amphorae = os_api_get_amphorae(auth, loadbalancer_url)

    logmsg("Get balancers")
    balancers = os_api_get_loadbalancers_with_status(auth, loadbalancer_url)

    logmsg("Get ports")
    ports = OsPorts(auth, network_url)

    report_file = get_report_file_name()
    report_dir = get_report_dir_name()

    logmsg("Prepare report")
    report_context = {
        "base_font": None,
        "header_font": None,
        "base_fill": None,
        "medium_fill": None,
        "high_fill": None,
        "sheets": dict()
    }
    header = ("id", "name", "provisioning_status", "operating_status", "age_since_create", "age_since_update", "amphorae", "members")
    wb = Workbook()
    ws = None
    for row in balancers.values():
        project = projects.get(row["details"]["project_id"], {"name": "__nonexistent__"})["name"]
        if not report_context["sheets"]:
            ws = wb.active
            ws.title = project
        if project not in wb.sheetnames:
            sheetnames = wb.sheetnames
            sheetnames.append(project)
            sheetnames.sort(key=str.lower)
            wb.create_sheet(project, sheetnames.index(project))
        ws = wb[project]
        if ws.title not in report_context["sheets"]:
            report_context["sheets"][ws.title] = dict()
            ws.append(header)
            report_context["sheets"][ws.title]["row_count"] = 1
            adjust_col_width(ws, report_context["sheets"][ws.title]["row_count"], header)
            if not report_context["base_font"]:
                report_context["base_font"] = copy(ws[1][0].font)
                report_context["header_font"] = copy(report_context["base_font"])
                report_context["header_font"].bold = True
                report_context["base_fill"] = copy(ws[1][0].fill)
                report_context["medium_fill"] = copy(report_context["base_fill"])
                report_context["high_fill"] = copy(report_context["base_fill"])
                report_context["medium_fill"].patternType = "solid"
                report_context["medium_fill"].fgColor = "00FFFF00"
                report_context["high_fill"].patternType = "solid"
                report_context["high_fill"].fgColor = "00FF0000"
            for cell in ws[report_context["sheets"][ws.title]["row_count"]]:
                cell.font = report_context["header_font"]
        row_data = {header.index(k) + 1: v for k, v in row["details"].items()
                    if k in ("id", "name", "provisioning_status", "operating_status")}
        row_data[header.index("age_since_create") + 1] = os_timestamp_age(row["details"]["created_at"])
        row_data[header.index("age_since_update") + 1] = os_timestamp_age(row["details"]["updated_at"] or row["details"]["created_at"])
        inspection_result = {
            "id": inspect_loadbalancer(row, ports),
            "amphorae": inspect_amphorae(amphorae.get(row["details"]["id"], tuple()), ports),
            "members": inspect_members(row, ports)
        }
        for k, v in inspection_result.items():
            if k == "id":
                continue
            row_data[header.index(k) + 1] = v["point"]
        ws.append(row_data)
        report_context["sheets"][ws.title]["row_count"] += 1
        for k, v in zip(("provisioning_status", "operating_status"), ("ACTIVE", "ONLINE")):
            if row["details"][k] == "ERROR":
                ws[report_context["sheets"][ws.title]["row_count"]][header.index(k)].fill = report_context["high_fill"]
            elif row["details"][k] != v:
                ws[report_context["sheets"][ws.title]["row_count"]][header.index(k)].fill = report_context["medium_fill"]
        for k, v in inspection_result.items():
            if v["severity"] == InspectionSeverity.HIGH:
                ws[report_context["sheets"][ws.title]["row_count"]][header.index(k)].fill = report_context["high_fill"]
            if v["severity"] == InspectionSeverity.MEDIUM:
                ws[report_context["sheets"][ws.title]["row_count"]][header.index(k)].fill = report_context["medium_fill"]
            ws[report_context["sheets"][ws.title]["row_count"]][header.index(k)].comment = Comment(v["comment"], "", width=288)
        adjust_col_width(ws, report_context["sheets"][ws.title]["row_count"], header)

    if not exists(report_dir):
        logmsg("Create report dir")
        os.mkdir(report_dir)
    else:
        logmsg("Clean the report dir up")
        for i in os.scandir(report_dir):
            if i.name.lower().endswith(".xlsx"):
                i_path = path_join(report_dir, i.name)
                logmsg("  Remove {}".format(i_path))
                os.remove(i_path)
    report_file_path = path_join(report_dir, report_file)
    logmsg("Write general report {}".format(report_file_path))
    wb.save(report_file_path)
    logmsg("Write per-project reports")
    orig_sheets = wb._sheets    # manipulate protected member for efficiency
    for sheetname in wb.sheetnames:
        wb._sheets = [sheet for sheet in orig_sheets if sheet.title == sheetname]
        report_file_path = path_join(report_dir, "project_{}.xlsx".format(sheetname))
        logmsg("  Write {}".format(report_file_path))
        wb.save(report_file_path)
