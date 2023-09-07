
# common lib
from report import *

# __main__-related imports
import json
from openpyxl import Workbook
# from openpyxl.comments import Comment
from copy import copy
from os.path import exists
from os.path import join as path_join


def os_api_get_volumes(auth_obj: OsAuth, url, page_sz=0, num_pages=0):
    return os_api_get_resources_gen(auth_obj, url, "volumes/detail", "volumes", None, page_sz, num_pages)


class OsVolumes:
    def __init__(self, auth_obj: OsAuth, url, users: OsUsers = None):
        self.auth = auth_obj
        self.url = url
        self.volumes = {"by_id": {}, "by_status": {}, "by_snapshot_id": {}}
        self.vol_count_by_status = dict()
        self.snapshots = {"all": [], "by_id": {}, "by_status": {}, "by_volume_id": {}}
        self.users = users
        self.refresh()

    def refresh(self, page_sz=0, num_pages=0):
        for i in os_api_get_resources_gen(self.auth, self.url, "snapshots/detail", "snapshots", "snapshots_links", page_sz, num_pages, {"all_tenants": True}):
            self.snapshots["all"].append(i)
            volume_id = i["volume_id"]
            status = i["status"]
            self.snapshots["by_id"][i["id"]] = i
            if volume_id not in self.snapshots["by_volume_id"]:
                self.snapshots["by_volume_id"][volume_id] = list()
            self.snapshots["by_volume_id"][volume_id].append(i)
            if status not in self.snapshots["by_status"]:
                self.snapshots["by_status"][status] = list()
            self.snapshots["by_status"][status].append(i)
        for i in os_api_get_resources_gen(self.auth, self.url, "volumes/detail", "volumes", "volumes_links", page_sz, num_pages, {"all_tenants": True}):
            if self.users:
                self.users.add_id(i["user_id"])
            self.volumes["by_id"][i["id"]] = i
            status = i["status"]
            snapshot_id = i["snapshot_id"]
            if status not in self.volumes["by_status"]:
                self.volumes["by_status"][status] = list()
            self.volumes["by_status"][status].append(i)
            if snapshot_id:
                if snapshot_id not in self.volumes["by_snapshot_id"]:
                    self.volumes["by_snapshot_id"][snapshot_id] = list()
                self.volumes["by_snapshot_id"][snapshot_id].append(i)
        self.vol_count_by_status = {x: len(self.volumes["by_status"][x]) for x in sorted(self.volumes["by_status"].keys())}

    def get_children_by_vol_id(self, vol_id):
        rv_volumes = list()
        children_ids = [vol_id]
        while children_ids:
            op_volumes = list()
            op_snapshots = list()
            for i in children_ids:
                op_snapshots.extend(self.snapshots["by_volume_id"].get(i, list()))
            for i in op_snapshots:
                op_volumes.extend(self.volumes["by_snapshot_id"].get(i["id"], list()))
            rv_volumes.extend(op_volumes)
            children_ids = list(x["id"] for x in op_volumes)
        return rv_volumes

    def get_children_by_snap_id(self, snap_id):
        op_volumes1 = self.volumes["by_snapshot_id"].get(snap_id, list())
        rv_volumes = op_volumes1
        while op_volumes1:
            op_volumes2 = list()
            for i in op_volumes1:
                op_volumes2.extend(self.get_children_by_vol_id(i["id"]))
            op_volumes1 = op_volumes2
            rv_volumes.extend(op_volumes1)
        return rv_volumes

    def get_by_id(self, vol_id):
        return self.volumes["by_id"].get(vol_id, dict())

    def get_vol_count_by_status(self):
        return self.vol_count_by_status

    def get_by_status(self, statuses: set):
        rv = list()
        for i in self.volumes["by_status"].keys():
            if i in statuses:
                rv.extend(self.volumes["by_status"][i])
        return rv

    def get_by_status_excluding(self, statuses: set):
        rv = list()
        for i in self.volumes["by_status"].keys():
            if i not in statuses:
                rv.extend(self.volumes["by_status"][i])
        return rv

    def get_snapshots(self):
        return self.snapshots["all"]

    def get_snapshot_by_id(self, snap_id):
        return self.snapshots["by_id"].get(snap_id, dict())


if __name__ == "__main__":
    logmsg("Auth")
    auth = OsAuth()

    users = OsUsers(auth, auth.identity_url)

    logmsg("Get projects")
    projects = os_api_get_projects(auth, auth.identity_url)

    logmsg("Get Volumes URL")
    volumes_url = os_api_get_service_url(auth, "volumev3")

    logmsg("Get volumes")
    volumes = OsVolumes(auth, volumes_url, users)

    logmsg("Get users")
    users.refresh()

    logmsg("Existing volumes statuses: {}".format(volumes.get_vol_count_by_status()))

    logmsg("Prepare report")
    report_file = get_report_file_name()
    report_dir = get_report_dir_name()
    report_context = {
        "base_font": None,
        "header_font": None,
        "base_fill": None,
        "medium_fill": None,
        "high_fill": None,
        "sheets": dict()
    }
    header = ("id", "name", "size", "status", "vol_user", "child_volumes", "age_since_create", "age_since_update")
    header_display = ("vol_id", "vol_name", "vol_size", "vol_status", "vol_user", "child_volumes", "vol_age_since_create", "vol_age_since_update")
    wb = Workbook()
    ws = None
    for row in volumes.get_by_status_excluding({"in-use"}):
        project = projects.get(row["os-vol-tenant-attr:tenant_id"], {"name": "__nonexistent__"})["name"]
        if not report_context["sheets"]:
            ws = wb.active
            ws.title = project
        if project not in wb.sheetnames:
            sheetnames = wb.sheetnames
            sheetnames.append(project)
            sheetnames.sort(key=str.lower)
            wb.create_sheet(project, sheetnames.index(project))
        ws = wb[project]
        if ws.title not in report_context["sheets"]:
            report_context["sheets"][ws.title] = dict()
            ws.append(header_display)
            report_context["sheets"][ws.title]["row_count"] = 1
            adjust_col_width(ws, report_context["sheets"][ws.title]["row_count"], header)
            if not report_context["base_font"]:
                report_context["base_font"] = copy(ws[1][0].font)
                report_context["header_font"] = copy(report_context["base_font"])
                report_context["header_font"].bold = True
                report_context["base_fill"] = copy(ws[1][0].fill)
                report_context["medium_fill"] = copy(report_context["base_fill"])
                report_context["high_fill"] = copy(report_context["base_fill"])
                report_context["medium_fill"].patternType = "solid"
                report_context["medium_fill"].fgColor = "00FFFF00"
                report_context["high_fill"].patternType = "solid"
                report_context["high_fill"].fgColor = "00FF0000"
            for cell in ws[report_context["sheets"][ws.title]["row_count"]]:
                cell.font = report_context["header_font"]
        row_data = {header.index(k) + 1: v for k, v in row.items() if k in ("id", "name", "size", "status")}
        row_data[header.index("vol_user") + 1] = users.get_by_id(volumes.get_by_id(row["id"])["user_id"]).get("name", "")
        row_data[header.index("child_volumes") + 1] = "\n".join("{} ({} in {})".format(x["id"], x["status"], projects.get(x["os-vol-tenant-attr:tenant_id"], {"name": "__nonexistent__"})["name"]) for x in volumes.get_children_by_vol_id(row["id"]))
        row_data[header.index("age_since_create") + 1] = os_timestamp_age(row["created_at"])
        row_data[header.index("age_since_update") + 1] = os_timestamp_age(row["updated_at"] or row["created_at"])
        ws.append(row_data)
        report_context["sheets"][ws.title]["row_count"] += 1
        adjust_col_width(ws, report_context["sheets"][ws.title]["row_count"], header)

    if not exists(report_dir):
        logmsg("Create report dir")
        os.mkdir(report_dir)
    else:
        logmsg("Clean the report dir up")
        for i in os.scandir(report_dir):
            if i.name.lower().endswith(".xlsx"):
                i_path = path_join(report_dir, i.name)
                logmsg("  Remove {}".format(i_path))
                os.remove(i_path)
    report_file_path = path_join(report_dir, report_file)
    logmsg("Write general report {}".format(report_file_path))
    wb.save(report_file_path)
    logmsg("Write per-project reports")
    orig_sheets = wb._sheets    # manipulate protected member for efficiency
    for sheetname in wb.sheetnames:
        wb._sheets = [sheet for sheet in orig_sheets if sheet.title == sheetname]
        report_file_path = path_join(report_dir, "project_{}.xlsx".format(sheetname))
        logmsg("  Write {}".format(report_file_path))
        wb.save(report_file_path)
