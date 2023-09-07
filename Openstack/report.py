
import requests
import os
import sys
from time import sleep
from urllib.parse import urlsplit, urlunsplit
from datetime import datetime
from collections import OrderedDict, namedtuple
from openpyxl.utils import get_column_letter
from os.path import basename


def os_timestamp_age(timestr):
    """
    Returns human-readable diff in utcnow and timestr
    :param timestr: Timestring given in ISO 8601 format ("%Y-%m-%dT%H:%M:%S") or ("%Y-%m-%dT%H:%M:%S.%f")
    :return: str
    """
    try:    # For old Python without fromisoformat method
        return str(datetime.utcnow().replace(microsecond=0) - datetime.strptime(timestr, "%Y-%m-%dT%H:%M:%S"))
    except ValueError:
        return str(datetime.utcnow().replace(microsecond=0) - datetime.strptime(timestr, "%Y-%m-%dT%H:%M:%S.%f").replace(microsecond=0))


def logmsg(msg):
    print(msg, file=sys.stderr)


def get_report_file_name():
    file_name = basename(sys.argv[0]).rsplit(".", 1)
    if len(file_name) > 1 and file_name[-1].strip().lower() == "py":
        file_name[-1] = "xlsx"
    else:
        file_name.append("xlsx")
    return  ".".join(file_name)


def get_report_dir_name():
    dir_name = basename(sys.argv[0]).rsplit(".", 1)
    if len(dir_name) > 1 and dir_name[-1].strip().lower() == "py":
        dir_name[-1] = "out"
    else:
        dir_name.append("out")
    return  "_".join(dir_name)


def adjust_col_width(sheet, row_i, header):
    for col_i in range(len(header)):
        value_len = len(str(sheet[row_i][col_i].value))
        if sheet.column_dimensions[get_column_letter(col_i + 1)].width <= value_len:
            sheet.column_dimensions[get_column_letter(col_i + 1)].width = (value_len + 1)


class OsAuth:
    def __init__(self):
        self.os_auth_obj = {
            "auth": {
                "identity": {
                    "methods": ["password"],
                    "password": {
                        "user": {
                            "name": os.environ["OS_USERNAME"],
                            "domain": {
                                "name": os.environ["OS_USER_DOMAIN_NAME"]
                            },
                            "password": os.environ["OS_PASSWORD"]
                        }
                    }
                },
                "scope": {
                    "project": {
                        "domain": {
                            "id": os.environ.get("OS_PROJECT_DOMAIN_ID", ""),
                            "name": os.environ.get("OS_PROJECT_DOMAIN_NAME", "")
                        },
                        "id": os.environ.get("OS_PROJECT_ID", ""),
                        "name": os.environ.get("OS_PROJECT_NAME", "")
                    }
                }
            }
        }
        self.os_auth_params = {
            # "nocatalog": True
        }
        self.r = requests.Response()
        self.catalog = dict()
        self.identity_url = ""
        self.refresh()

    def get_token(self):
        return self.r.headers["X-Subject-Token"]

    def get_data(self):
        return self.r.json()

    def get_endpoint_url(self, service):
        return self.catalog[service][os.environ["OS_REGION_NAME"]][os.environ["OS_INTERFACE"]][0]

    def find_identity_url(self):
        r = requests.get(os.environ["OS_AUTH_URL"])
        r.raise_for_status()
        if "version" in r.json():   # we're at the real endpoint
            for link in r.json()["version"]["links"]:
                if link["rel"] == "self":
                    self.identity_url = link["href"]
                    break
        elif "versions" in r.json(): # we're dealing with a catalogue
            links = OrderedDict.fromkeys(("current", "stable"))
            for version in r.json()["versions"]["values"]:
                status = version["status"].lower()
                if status in links.keys():
                    for link in version["links"]:
                        if link["rel"] == "self":
                            links[status] = link["href"]
                            break
            for status in links.keys():
                if links[status]:
                    self.identity_url = links[status]
                    break

    def refresh(self):
        self.find_identity_url()
        self.r = requests.post("/".join((self.identity_url, "auth/tokens")), json=self.os_auth_obj, params=self.os_auth_params)
        self.r.raise_for_status()
        self.refresh_catalog()

    # Catalog:
    # Service -> Region -> Interface -> URL, URL, URL...
    #                   -> Interface -> URL, URL, URL...
    #                   -> Interface -> URL, URL, URL...
    #            Region -> Interface -> URL, URL, URL...
    #                   -> Interface -> URL, URL, URL...
    #                   -> Interface -> URL, URL, URL...
    def refresh_catalog(self):
        for service in self.r.json()["token"]["catalog"]:
            service_type = service["type"]
            if service_type not in self.catalog:
                self.catalog[service_type] = dict()
            for endpoint in service["endpoints"]:
                endpoint_region = endpoint["region"]
                endpoint_interface = endpoint["interface"]
                endpoint_url = endpoint["url"]
                if endpoint_region not in self.catalog[service_type]:
                    self.catalog[service_type][endpoint_region] = dict()
                if endpoint_interface not in self.catalog[service_type][endpoint_region]:
                    self.catalog[service_type][endpoint_region][endpoint_interface] = list()
                self.catalog[service_type][endpoint_region][endpoint_interface].append(endpoint_url.rstrip("/"))


def os_api_get(auth_obj: OsAuth, url=None, path="", service="", params=None, auth_tries_max=3, sleep_before_retry=1):
    if not params:
        params = dict()
    r = requests.Response()
    success = False
    auth_tries = 0
    while not success:
        try:
            url = url or auth_obj.get_endpoint_url(service)
            url = "/".join((url, path)) if path else url
            r = requests.get(url, headers={"X-Auth-Token": auth_obj.get_token()}, params=params, cert=os.environ.get("OS_CACERT", None))
            r.raise_for_status()
            success = True
        except requests.HTTPError as e:
            if e.response.status_code == 401:
                if auth_tries < auth_tries_max:
                    auth_tries += 1
                    print("Re-trying auth", file=sys.stderr)
                    sleep(sleep_before_retry)
                    auth_obj.refresh()
                    pass
            else:
                raise
    return r


def os_api_get_service_url(auth_obj: OsAuth, service):
    endpoint_url = auth_obj.get_endpoint_url(service)
    url_scheme = urlsplit(endpoint_url)[0]
    try:
        for version in os_api_get(auth_obj, service=service).json()["versions"]:
            if version["status"] == "CURRENT":
                for link in version["links"]:
                    url_parts = list(urlsplit(link["href"]))
                    url_parts[0] = url_scheme
                    if link["rel"] == "self":
                        return urlunsplit(url_parts)
    except requests.HTTPError:  # version discovery cannot be performed on endpoint_url
        return endpoint_url


def os_api_get_resources_gen(auth_obj: OsAuth, url, path, items_key, links_key=None, page_sz=0, num_pages=0, add_params:dict=None):
    """
    Creates a generator over a list of openstack resources of a kind
    :param auth_obj: Openstack auth obj
    :param url: Explicit URL
    :param path: API path
    :param items_key: Items key name in the returned object
    :param links_key: Links key name in the returned object (None means no pagination assumed/supported)
    :param page_sz: Result pagination limit (positive integer or 0 to disable limit)
    :param num_pages: Number of pages to request (positive integer or 0 to disable limit)
    :param add_params: Additional non-generic params that have impact on data retrieval
    :return: Generator over openstack resources of a kind
    """
    params = add_params or dict()
    if page_sz > 0:
        params["limit"] = page_sz
    url_scheme = urlsplit(url)[0]
    r = os_api_get(auth_obj, url, path, params=params)
    page_count = 0
    while True:
        yield from r.json()[items_key]
        page_count += 1
        if links_key and links_key in r.json() and (num_pages == 0 or (num_pages > 0 and page_count < num_pages)):
            url = None
            for link in r.json()[links_key]:
                if link["rel"] == "next":
                    url_parts = list(urlsplit(link["href"]))
                    url_parts[0] = url_scheme
                    url = urlunsplit(url_parts)
                    break
            if url:
                r = os_api_get(auth_obj, url)
            else:
                break
        else:
            break


def os_api_get_projects(auth_obj: OsAuth, url):
    """
    Returns proejcts as a dict {project_id -> project_details...}
    :param auth_obj: Openstack auth obj
    :param url: Explicit URL
    :return: Dict of projects
    """
    rv = dict()
    for i in os_api_get_resources_gen(auth_obj, url, "projects", "projects"):
        rv[i["id"]] = i
    return rv


def os_api_get_ports(auth_obj: OsAuth, url, page_sz=0, num_pages=0):
    rv = dict()
    for i in os_api_get_resources_gen(auth_obj, url, "ports", "ports", "ports_links", page_sz, num_pages):
        rv[i["id"]] = i
    return rv


class OsPorts:
    def __init__(self, auth_obj:OsAuth, url):
        self.auth = auth_obj
        self.url = url
        self.ports = dict()
        self.refresh()

    def refresh(self, page_sz=0, num_pages=0):
        self.ports = {"by_id": {}, "by_subnet_ip": {}, "by_network": {}}
        for i in os_api_get_resources_gen(self.auth, self.url, "ports", "ports", "ports_links", page_sz, num_pages):
            self.ports["by_id"][i["id"]] = i
            if i["network_id"] not in self.ports["by_network"]:
                self.ports["by_network"][i["network_id"]] = list()
            self.ports["by_network"][i["network_id"]].append(i)
            for j in i["fixed_ips"]:
                self.ports["by_subnet_ip"]["_".join((j["subnet_id"], j["ip_address"]))] = i

    def get_by_id(self, port_id):
        return self.ports["by_id"].get(port_id, dict())

    def get_by_subnet_ip(self, subnet, ip):
        return self.ports["by_subnet_ip"].get("_".join((subnet, ip)), dict())

    def get_by_network(self, network_id):
        return self.ports["by_network"].get(network_id, [])


class OsUsers:
    def __init__(self, auth_obj:OsAuth, url):
        self.auth = auth_obj
        self.url = url
        self.user_ids = set()
        self.users = dict()

    def add_id(self, user_id):
        self.user_ids.add(user_id)

    def refresh(self):
        self.users = {"by_id": {}}
        for i in self.user_ids:
            self.users["by_id"][i] = os_api_get(self.auth, self.url, "users/{}".format(i)).json()["user"]
        self.user_ids = set()

    def get_by_id(self, user_id):
        return self.users["by_id"].get(user_id, dict())


InspectionSeverity = namedtuple("InspectionSeverity", ("LOW", "MEDIUM", "HIGH"))(*range(1, 4))


def make_inspection_rv(inspections, rv):
    inspection = dict()
    for k, v in inspections.items():
        if v["count"]:
            inspection["point"] = k
            inspection["severity"] = v["severity"]
            break
    comment = "\n".join(rv)
    return {"point": inspection["point"], "severity": inspection["severity"], "comment": comment}
